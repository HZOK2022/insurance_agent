# -*- coding: utf-8 -*-
"""保费计算(PremiumStore 费率事实源 + calculate_premium 工具)。

设计(定稿):
- products: id 自增主键;key 稳定业务键/代码(LLM 引用+外键);name 显示名;kb_doc_id 关联向量库条款文档。
- premium_rates: id 自增主键;product_key 外键→products.key;无 chunk_id(费率行不是文档片段);
  引用角标直接用该行 id(快照在会话 retrieval 事件里,事件解析可溯)。
- 不同产品计算方式/附加包不同 → calculate_premium 按 product_key 分发到各产品计算器(策略)。
- 金额查表+确定计算(重疾每5万保额、家庭单折扣),LLM 不手算。
"""
from __future__ import annotations

import json
import os
import re
import sqlite3
from typing import Any

PRODUCT_XX = "尊享e生2025"                        # key(稳定业务键,LLM 引用)
PRODUCT_XX_NAME = "尊享 e 生·中高端医疗保险 PLUS（2025版）（年缴版）"
VERSION_XX = "v2025"

_XX_PLAN_COLS = [
    ("0元", "计划一"), ("0元", "计划二"),
    ("1.5万", "计划一"), ("1.5万", "计划二"),
    ("3万", "计划一"), ("3万", "计划二"),
]
_XX_ADDON_COLS = [
    ("family_deductible", "家庭共享免赔额", {}, "元/年"),
    ("clinic_a", "门急诊加油包A-不含器质", {}, "元/年"),
    ("clinic_b", "门急诊加油包B-含器质", {}, "元/年"),
    ("drug", "药费院加油包", {}, "元/年"),
    ("critical", "重疾加油包（每5万保额）", {"gender": "男"}, "元/每5万保额"),
    ("critical", "重疾加油包（每5万保额）", {"gender": "女"}, "元/每5万保额"),
]


def _age_range(label):
    m = re.match(r"\[\s*(\d+)\s*,\s*(\d+)\s*\]", label.strip())
    if not m:
        raise ValueError(f"bad age label: {label!r}")
    return int(m.group(1)), int(m.group(2))


def _num(val):
    if val is None:
        return None
    s = str(val).strip().replace(",", "")
    if s == "" or s in ("不适用", "不适用。", "N/A", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _fmt_dims(dims):
    if not dims:
        return ""
    return "(" + ",".join(f"{v}" for v in dims.values()) + ")"


class PremiumStore:
    """费率事实源(SQLite)。单写者;只 INSERT/UPSERT。"""

    def __init__(self, path):
        d = os.path.dirname(path)
        if d:
            os.makedirs(d, exist_ok=True)
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA journal_mode=WAL")
        self._init_schema()

    def _init_schema(self):
        self.conn.executescript("""
        CREATE TABLE IF NOT EXISTS products (
          id         INTEGER PRIMARY KEY AUTOINCREMENT,
          key        TEXT NOT NULL UNIQUE,
          name       TEXT NOT NULL,
          kb_doc_id  TEXT,
          version    TEXT,
          coverage   TEXT,
          rules      TEXT,
          calc_config TEXT,
          source     TEXT
        );
        CREATE TABLE IF NOT EXISTS premium_rates (
          id          INTEGER PRIMARY KEY AUTOINCREMENT,
          product_key TEXT NOT NULL,
          item_key    TEXT NOT NULL,
          item_name   TEXT,
          dims        TEXT NOT NULL DEFAULT '{}',
          age_min     INTEGER NOT NULL,
          age_max     INTEGER NOT NULL,
          premium     REAL,
          unit        TEXT,
          source      TEXT,
          section     TEXT,
          UNIQUE(product_key, item_key, dims, age_min, age_max)
        );
        CREATE INDEX IF NOT EXISTS idx_rate_lookup ON premium_rates(product_key, item_key, age_min, age_max);
        """)
        self.conn.commit()

    def upsert_product(self, key, name, kb_doc_id, version, coverage, rules, calc_config, source):
        self.conn.execute(
            """INSERT INTO products(key, name, kb_doc_id, version, coverage, rules, calc_config, source)
               VALUES(?,?,?,?,?,?,?,?)
               ON CONFLICT(key) DO UPDATE SET name=excluded.name, kb_doc_id=excluded.kb_doc_id,
                 coverage=excluded.coverage, rules=excluded.rules, calc_config=excluded.calc_config""",
            (key, name, kb_doc_id, version, coverage, rules,
             json.dumps(calc_config, ensure_ascii=False).replace("\n", ""), source))
        self.conn.commit()

    def upsert_rate(self, product_key, item_key, item_name, dims, age_min, age_max,
                    premium, unit, source, section):
        d = json.dumps(dims, sort_keys=True, ensure_ascii=False)
        self.conn.execute(
            """INSERT INTO premium_rates(product_key, item_key, item_name, dims, age_min, age_max,
                 premium, unit, source, section)
               VALUES(?,?,?,?,?,?,?,?,?,?)
               ON CONFLICT(product_key, item_key, dims, age_min, age_max)
               DO UPDATE SET premium=excluded.premium, item_name=excluded.item_name, unit=excluded.unit""",
            (product_key, item_key, item_name, d, age_min, age_max, premium, unit, source, section))
        self.conn.commit()
        row = self.conn.execute(
            """SELECT * FROM premium_rates WHERE product_key=? AND item_key=? AND dims=? AND age_min=? AND age_max=?""",
            (product_key, item_key, d, age_min, age_max)).fetchone()
        out = dict(row)
        out["dims"] = json.loads(out["dims"])
        return out

    def get_rate(self, product_key, item_key, dims, age):
        cur = self.conn.execute(
            "SELECT * FROM premium_rates WHERE product_key=? AND item_key=? AND ?>=age_min AND ?<=age_max",
            (product_key, item_key, age, age))
        for row in cur.fetchall():
            if json.loads(row["dims"]) == dims:
                d = dict(row)
                d["dims"] = json.loads(d["dims"])
                return d
        return None

    def get_product(self, ref):
        """按 key 或 name 查产品。"""
        ref = (ref or "").strip()
        if not ref:
            return None
        row = self.conn.execute("SELECT * FROM products WHERE key=? OR name=?", (ref, ref)).fetchone()
        if not row:
            return None
        d = dict(row)
        d["calc_config"] = json.loads(d["calc_config"]) if d.get("calc_config") else {}
        return d

    def close(self):
        self.conn.close()


CALCULATORS = {}


def register(product_key):
    def deco(fn):
        CALCULATORS[product_key] = fn
        return fn
    return deco


@register(PRODUCT_XX)
def _calc_xx2025(store, product_key, age, items, family_member_count=1):
    lines, rows, total = [], [], 0.0
    for it in items:
        item_key = it.get("item_key")
        dims = it.get("dims") or {}
        coverage = it.get("coverage")
        row = store.get_rate(product_key, item_key, dims, age)
        if not row:
            lines.append(f"方案 {item_key}{_fmt_dims(dims)}:未找到费率")
            continue
        if row["premium"] is None:
            lines.append(f"{row['item_name']}{_fmt_dims(dims)}:该年龄段不适用")
            rows.append(row)
            continue
        amount = float(row["premium"])
        if row.get("unit") == "元/每5万保额" and coverage:
            amount = amount * (coverage / 50000)
        lines.append(f"{row['item_name']}{_fmt_dims(dims)}: {amount:,.2f}元/年 [{len(rows)+1}]")
        total += amount
        rows.append(row)
    disc = 1.0
    if family_member_count is not None and family_member_count >= 3:
        disc = 0.90
    elif family_member_count is not None and family_member_count == 2:
        disc = 0.95
    total_x = total * disc
    disc_note = "" if disc == 1.0 else f" (家庭单优享 {family_member_count}人 → {disc*100:.0f}折)"
    header = f"{PRODUCT_XX_NAME} {VERSION_XX} 保费测算(年龄 {age}, 家庭单 {family_member_count or 1} 人):"
    content = f"{header}\n" + "\n".join(lines) + f"\n合计: {total_x:,.2f}元/年{disc_note}"
    return content, rows


def calculate_premium(store, args):
    ref_name = (args.get("product") or "").strip()
    if not ref_name:
        return {"content": "请指定产品(product,填产品 key 或名称)", "reference": []}
    prod = store.get_product(ref_name)
    if not prod:
        return {"content": f"产品 {ref_name} 不存在/暂无费率", "reference": []}
    key = prod["key"]
    calc = CALCULATORS.get(key)
    if not calc:
        return {"content": f"产品 {ref_name} 暂无保费计算配置", "reference": []}
    try:
        age = int(args.get("age"))
    except (TypeError, ValueError):
        return {"content": "年龄(age)无效", "reference": []}
    items = args.get("items") or []
    if not isinstance(items, list) or not items:
        return {"content": "请至少指定一个方案(items)", "reference": []}
    try:
        family = int(args.get("family_member_count")) if args.get("family_member_count") is not None else 1
    except (TypeError, ValueError):
        family = 1
    content, store_rows = calc(store, key, age, items, family)
    ref = []
    for r in store_rows:
        ref.append({
            "chunk_id": str(r["id"]), "score": None,
            "doc_id": key, "version": prod.get("version", ""),
            "section": r.get("section") or r.get("item_name", ""),
            "source": r.get("source", ""),
            "content": f"{r['item_name']}{_fmt_dims(r['dims'])}: {r['premium']}元/年({r['unit']})",
        })
    return {"content": content, "reference": ref}


def build_premium_tool(store):
    schema = {"type": "function", "function": {
        "name": "calculate_premium",
        "description": "按产品/投保年龄/方案计算年缴保费(查表确定性)。items=[{item_key,dims?,coverage?}]。item_key/dims:尊享e生2025→必选 plan:{deductible:'0元|1.5万|3万',plan_variant:'计划一|计划二'};加油包 family_deductible/clinic_a/clinic_b/drug:{},critical:{gender:'男|女'}且 coverage=保额(每5万保额);安盛天平卓越馨选2025→住院 hospital:{deductible:'0元|5000元|10000元|15000元|20000元',tier:'普A|普B|普C|特A|特B|特C',social:'有社保|无社保'},门急诊 outpatient:{deductible:'0元|200元|500元|1300元',coverage:'1万|1.5万|2万|3.5万',social:'有社保|无社保'},重疾津贴 majordaily:{version:'普通版|特需版',plan:'A|B|C'},重疾保险金 majorsum:{version:'普通版|特需版'},博鳌 boao:{type:'特药械|院外特定药品|特定医疗器械',social:'有社保|无社保|有/无社保'}。product 传 key 或名称(如 尊享e生2025 / 安盛天平卓越馨选2025);family_member_count≥2 享家庭单折扣。返回可读账单+可溯源引用。对比两方案/两口径(如 有/无社保)请调本工具分别算再比。",
        "parameters": {"type": "object", "properties": {
            "product": {"type": "string", "description": "产品 key 或名称"},
            "age": {"type": "integer", "description": "投保年龄(周岁)"},
            "items": {"type": "array", "description": "要计算的方案/包,可多选",
                      "items": {"type": "object", "properties": {
                          "item_key": {"type": "string"},
                          "dims": {"type": "object", "description": "按 description 中该 item_key 的 dims 构造(详见 description)"},
                          "coverage": {"type": "number", "description": "保额(元),仅 critical 需传"},
                      }, "required": ["item_key"]}},
            "family_member_count": {"type": "integer", "description": "家庭单成员数;2人95折,≥3人9折"},
        }, "required": ["product", "age", "items"]}}}

    def handler(args, start_idx=0):
        return calculate_premium(store, args or {})

    return {"schema": schema, "handler": handler}


def load_xx2025_xlsx(store, xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    n = 0
    source = xlsx_path
    ws = wb["必选计划费率表"]
    for row in ws.iter_rows(min_row=6, values_only=True):
        label = row[0]
        if label is None or not str(label).strip().startswith("["):
            continue
        age_min, age_max = _age_range(str(label))
        for i, (deductible, plan) in enumerate(_XX_PLAN_COLS):
            store.upsert_rate(product_key=PRODUCT_XX, item_key="plan",
                              item_name=f"必选计划({deductible}年免赔额,{plan})",
                              dims={"deductible": deductible, "plan_variant": plan},
                              age_min=age_min, age_max=age_max, premium=_num(row[i + 1]),
                              unit="元/年", source=source, section="必选计划费率表")
            n += 1
    ws = wb["加油包费率表"]
    for row in ws.iter_rows(min_row=6, values_only=True):
        label = row[0]
        if label is None or not str(label).strip().startswith("["):
            continue
        age_min, age_max = _age_range(str(label))
        for i, (item_key, item_name, dims, unit) in enumerate(_XX_ADDON_COLS):
            store.upsert_rate(product_key=PRODUCT_XX, item_key=item_key, item_name=item_name, dims=dims,
                              age_min=age_min, age_max=age_max, premium=_num(row[i + 1]),
                              unit=unit, source=source, section="加油包费率表")
            n += 1
    store.upsert_product(
        key=PRODUCT_XX, name=PRODUCT_XX_NAME, kb_doc_id=PRODUCT_XX, version=VERSION_XX,
        coverage=("一般医疗(对应年免赔额)+特定疾病医疗+外购药品及外购医疗器械费用医疗+特定药品费用医疗"
                  "+恶性肿瘤先进疗法医疗+特定疾病异地转诊公共交通费用及住宿费用+特定疾病住院津贴"
                  "+意外紧急牙齿门急诊医疗费用+全球紧急救援服务。"),
        rules=("首次投保年龄:出生满30天-70周岁;71周岁及以上仅限保单期满指定期限内重新投保;"
               "投保保单数量=1,个人单标准保费。家庭单优享:同投保人 尊享/众民保 系列家庭成员 2人享95折,3人及以上享9折。"),
        calc_config={"mandatory": ["plan"],
                     "optional": ["family_deductible", "clinic_a", "clinic_b", "drug", "critical"],
                     "unit_map": {"critical": 50000},
                     "discounts": {"family": [[2, 0.95], [3, 0.90]]},
                     "item_dims": {"plan": ["deductible", "plan_variant"], "critical": ["gender"]}},
        source=source)
    return n
